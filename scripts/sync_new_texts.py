"""
imports/texts/*.txt を確認し、imports/list.xlsx の「作品リスト」シートに
まだ登録されていないファイルの分だけ、行を自動で追加するスクリプト。

追加するのはファイル名(A列)とタイトル(B列、ファイル名を仮タイトルとして使用)のみ。
それ以外の列（日付・種別・ジャンル・カップリング・年齢指定・あらすじ・公開範囲）は
空欄のままにし、既存の行には一切手を加えない。

使い方:
    python scripts/sync_new_texts.py
"""
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "imports" / "list.xlsx"
TEXTS_DIR = ROOT / "imports" / "texts"


def main():
    if not XLSX_PATH.exists():
        print(f"[エラー] {XLSX_PATH} が見つかりません。imports/list.xlsx を用意してください。")
        sys.exit(1)

    if not TEXTS_DIR.exists():
        print(f"[エラー] {TEXTS_DIR} が見つかりません。")
        sys.exit(1)

    # 数式・書式を壊さないよう data_only は使わず読み込む（このあと上書き保存するため）
    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb["作品リスト"]

    existing_filenames = set()
    last_filled_row = 1  # ヘッダー行
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        filename = row[0]
        if filename is not None and str(filename).strip():
            existing_filenames.add(str(filename).strip())
            last_filled_row = row_idx

    txt_stems = sorted(p.stem for p in TEXTS_DIR.glob("*.txt"))
    new_stems = [stem for stem in txt_stems if stem not in existing_filenames]

    if not new_stems:
        print("追加対象のファイルはありませんでした（すべて登録済みです）。")
        return

    next_row = last_filled_row + 1
    for offset, stem in enumerate(new_stems):
        row_idx = next_row + offset
        ws.cell(row=row_idx, column=1).value = stem  # ファイル名
        ws.cell(row=row_idx, column=2).value = stem  # タイトル（仮）
        # 日付・種別・ジャンル・カップリング・年齢指定・あらすじ・公開範囲は空欄のまま

    wb.save(XLSX_PATH)

    print(f"追加: {len(new_stems)}件")
    for stem in new_stems:
        print(f" - {stem}")


if __name__ == "__main__":
    main()
